
import os
import shutil
from time import sleep
import openpyxl.worksheet.table
from openpyxl.utils import get_column_letter
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

reports_dir = "reports"
source_dir = "source"

try:
    shutil.rmtree(reports_dir)
except:
    print("error deleting reports directory structure")
    sleep(3)

# wait for windows to finish
sleep(3)
os.mkdir(reports_dir)

# styles
font_default = Font()
font_cap = Font(bold=True, color='FFFFFFFF')
fill_cap = PatternFill(fill_type='solid', start_color='2e6990', end_color='2e6990')
border = Border(
    left=Side(border_style='thin', color='FF000000'),
    right=Side(border_style='thin', color='FF000000'),
    top=Side(border_style='thin', color='FF000000'),
    bottom=Side(border_style='thin', color='FF000000'),
)
alignment=Alignment(horizontal='center')
number_format = 'mm/dd/yyyy'
protection = Protection()

for subdir, dirs, files in os.walk(source_dir):
    for filename in files:
        filepath = subdir + os.sep + filename
        workbook = openpyxl.load_workbook(filepath)

        for sheet in workbook.worksheets:

            # remove conditional formating
            sheet.conditional_formatting = ConditionalFormattingList()

            # unhide columns
            table_width = sheet.max_column
            for i in range(1, table_width):
                column_letter = get_column_letter(i)
                sheet.column_dimensions[column_letter].hidden = False

            # unhide rows and set row height to none
            table_length = sheet.max_row
            for i in range(0, table_length):
                sheet.row_dimensions[i].hidden = False
                sheet.row_dimensions[i].height = None

            # format for all cells
            for row in sheet.rows:
                for cell in row:
                    cell.font = font_default
                    cell.border = border
                    cell.alignment = alignment
                    cell.number_format = number_format
                    cell.protection = protection

            # format for first row
            first_row = list(sheet.rows)[0]
            for cell in first_row:
                cell.font = font_cap
                cell.fill = fill_cap

            # autofit column width
            for column in sheet.columns:
                length = 12  # as minimum
                for cell in column:
                    if len(str(cell.value)) > length:
                        length = len(str(cell.value))
                sheet.column_dimensions[column[1].column_letter].width = length

            sheet.column_dimensions[get_column_letter(table_width)].width = 12

        new_filepath = reports_dir + os.sep + filename
        workbook.save(new_filepath)
