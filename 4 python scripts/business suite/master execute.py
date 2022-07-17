
# create and set project directory
# set list of timezones needed
# devices set list of columns needed
# agents set list of columns needed
# audits set list of columns needed !!! important for every new type of column added in audits files
# audits set list of sheets to delete
# imacs set list of columns needed
# imacs rename any column label if needed

from asyncio.windows_events import NULL
import os
import shutil
from tkinter.font import names
from turtle import width
import dateutil
from time import sleep
import re
import csv
import pandas
from functools import reduce
from collections import OrderedDict
import openpyxl.worksheet.table
from openpyxl.styles import PatternFill, Alignment, NamedStyle, Font, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import pprint

# hardcode project location to circumvent windows error
os.chdir('C:/Users/mi250175/m_scripts')

# get configurations
from customers import masterdict
from customers import sort_order_ssq

# parse customer configuration
def parse_configuration(abbr_dict, temp_match_element, ThingName, Customer, HostName, data_dict, Installed):
    flag = 0
    for abbr_dict_key in abbr_dict:
        if flag == 1:
            break
        element_list = abbr_dict[abbr_dict_key]
        for element in element_list:
            if element == temp_match_element:
                data_dict_key = (ThingName, Customer, HostName)
                if data_dict_key not in data_dict:
                    data_dict[data_dict_key] = {}
                data_dict[data_dict_key][abbr_dict_key] = Installed
                flag = 1
                break

print("setting files & directories")

source_dir = "source"
source_dir_audits = source_dir + os.sep + "audits"
source_dir_imacs = source_dir + os.sep + "imacs"
source_dir_agents = source_dir + os.sep + "agents"
source_dir_devices = source_dir + os.sep + "devices"
source_dir_inventories = source_dir + os.sep + "inventories"

reports_dir = "reports"
reports_dir_primary = reports_dir + os.sep + "primary_reports"
reports_dir_single = reports_dir + os.sep + "single_customers"
reports_dir_audits = reports_dir + os.sep + "audit reports"

try:
    shutil.rmtree(reports_dir)
except:
    print("error deleting reports directory structure")
    sleep(3)
# wait for windows to finish
sleep(3)
os.mkdir(reports_dir)
os.mkdir(reports_dir_primary)
os.mkdir(reports_dir_single)
os.mkdir(reports_dir_audits)

# devices

zones = {
    "EST" : ["est", "EST", "Est", "-300.0", "-300"],
    "CST" : ["cst", "CST", "Cst", "CDT", "Central", "CNT", "-360.0", "-360"],
    "MST" : ["mst", "MST", "Mst", "-420.0", "-420"],
    "PST" : ["pst", "PST", "Pst", "PDT", "-480.0", "-480"],
    "Germany" : ["CEST UTC +2", "60.0", "60"]
}

print("parsing through devices files")
list_of_devices_dataframes = []
for file in os.scandir(source_dir_devices):
    # print("parsing through file ", file.name)
    dataframe = pandas.read_csv(file)
    list_of_devices_dataframes.append(dataframe)
master_devices_dataframe = pandas.concat(list_of_devices_dataframes, ignore_index=True)
list_of_columns_in_dataframe_all = list(master_devices_dataframe.columns)
list_of_columns_in_dataframe_needed = [
    "Customer",
    "name",
    "Asset_Name",
    "Asset_IP",
    "timeZoneOffset",
    "biosVersion",
    "isConnected",
    "Asset_DCSAgentVersion",
    # "BSBProduct" # estoril etc
]
list_of_columns_in_dataframe_to_drop = []
for item in list_of_columns_in_dataframe_all:
    if item not in list_of_columns_in_dataframe_needed:
        list_of_columns_in_dataframe_to_drop.append(item)
for column in list_of_columns_in_dataframe_to_drop:
    master_devices_dataframe.drop(column, axis='columns', inplace=True)
# label index column
master_devices_dataframe.index.names = ['index']
# filter out customers
master_devices_dataframe = master_devices_dataframe[master_devices_dataframe["Customer"].isin(masterdict)]
# order columns
master_devices_dataframe = master_devices_dataframe[list_of_columns_in_dataframe_needed]
# rename columns
master_devices_dataframe.rename(columns={"Asset_Name": "ATM ID"}, inplace=True)
master_devices_dataframe.rename(columns={"Customer": "Customer from devices"}, inplace=True)
# all ids to upper
master_devices_dataframe['ATM ID'] = master_devices_dataframe['ATM ID'].apply(lambda x: x.upper())
# cast timezone to string
master_devices_dataframe['timeZoneOffset'] = master_devices_dataframe['timeZoneOffset'].astype(str)
# unify timezone
for index, row in master_devices_dataframe.iterrows():
    for zone in zones:
        for element in zones[zone]:
            if element == row['timeZoneOffset']:
                master_devices_dataframe.at[index, 'timeZoneOffset'] = zone

# agents

print("parsing through agents files")
list_of_agents_dataframes = []
for file in os.scandir(source_dir_agents):
    # print("parsing through file ", file.name)
    dataframe = pandas.read_csv(file)
    list_of_agents_dataframes.append(dataframe)
master_agents_dataframe = pandas.concat(list_of_agents_dataframes, ignore_index=True)
list_of_columns_in_dataframe_all = list(master_agents_dataframe.columns)
list_of_columns_in_dataframe_needed = [
    "customer",
    # "name",
    "hostName",
    # "isConnected",
    # "dcsAgentVersion",
    "osVersion"
]
list_of_columns_in_dataframe_to_drop = []
for item in list_of_columns_in_dataframe_all:
    if item not in list_of_columns_in_dataframe_needed:
        list_of_columns_in_dataframe_to_drop.append(item)
for column in list_of_columns_in_dataframe_to_drop:
    master_agents_dataframe.drop(column, axis='columns', inplace=True)
# label index column
master_agents_dataframe.index.names = ['index']
# filter out customers
master_agents_dataframe = master_agents_dataframe[master_agents_dataframe["customer"].isin(masterdict)]
# order columns
master_agents_dataframe = master_agents_dataframe[list_of_columns_in_dataframe_needed]
# rename columns
master_agents_dataframe.rename(columns={"hostName": "ATM ID"}, inplace=True)
master_agents_dataframe.rename(columns={"customer": "Customer from agents"}, inplace=True)
# put all ids to upper
master_agents_dataframe['ATM ID'] = master_agents_dataframe['ATM ID'].apply(lambda x: x.upper())

# inventory

# getting customers configuration
ordered_config = masterdict
# add index for sorting configuration file
indexed_config = {}
index = 1
# only customers from parsed source data
customers_list_active = set()
for customer_key in ordered_config:
    customer_dict = ordered_config[customer_key]
    # adding indexed dictionary to customer
    indexed_config[customer_key] = {}
    for data_type in customer_dict:
        if data_type == "package data":
            packages_dict = customer_dict[data_type]
            for package_key in packages_dict:
                new_package_key = (index, package_key)
                package_value = packages_dict[package_key]
                indexed_config[customer_key][new_package_key] = package_value
                index += 1

##############################
# data structure for receiving data from csv file

# dictionary/list combo data structure: {
#   tuple(name1, name2, customer) : {software : date, software : date, ...}
# }
data_dict = {}
# dictionary/list combo data structure with added index: {
#   tuple(name1, name2, customer) : {tuple(index, software) : date, tuple(index, software) : date, ...}
# }
##############################

print("parsing through inventory data")
log_line_count = 0
for file in os.scandir(source_dir_inventories):
    # print("parsing through file: " + file.name)
    with open(file) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        for row in csv_reader:
            if row[0] == "ThingName":
                continue
            else:
                Customer = row[2]
                if Customer in ordered_config:
                    # add customer in set of active customers
                    customers_list_active.add(Customer)

                    ThingName = row[0]
                    HostName = row[1]
                    Version = row[4]

                    Installed = row[5]
                    try:
                        Installed = dateutil.parser.parse(Installed)
                        Installed = Installed.date()
                    except:
                        pass

                    # override for dcs, core/icusn and winpass
                    # put version in field instead of date for dcs & core
                    # put date back in the field instead of empty for winpass
                    DetectedSoftware = row[3]

                    regex_pattern_object = re.compile("NCR DCS Agent")
                    match_object = re.search(regex_pattern_object,DetectedSoftware)
                    if match_object:
                        Installed = Version

                    regex_pattern_object = re.compile("Edge-StandardInterface-Core")
                    match_object = re.search(regex_pattern_object,DetectedSoftware)
                    if match_object:
                        Installed = Version

                    regex_pattern_object = re.compile("ICUSN-Base")
                    match_object = re.search(regex_pattern_object,DetectedSoftware)
                    if match_object:
                        Installed = Version

                    regex_pattern_object = re.compile("ICUSN-MUP")
                    match_object = re.search(regex_pattern_object,DetectedSoftware)
                    if match_object:
                        Installed = Version

                    regex_pattern_object = re.compile("AE")
                    match_object = re.search(regex_pattern_object,DetectedSoftware)
                    if match_object:
                        Installed = Version

                    regex_pattern_object = re.compile("WinPass change date")
                    match_object = re.search(regex_pattern_object,DetectedSoftware)
                    if match_object:
                        DetectedSoftware = DetectedSoftware.replace("WinPass change date", "")
                        Installed = dateutil.parser.parse(DetectedSoftware)
                        Installed = Installed.date()
                        DetectedSoftware = "WinPass change date"
                        Version = "1.0"

                    # element to match
                    temp_match_element = [DetectedSoftware, Version]

                    # parse specific customer
                    abbr_dict_specific = indexed_config[Customer]
                    parse_configuration(abbr_dict_specific, temp_match_element, ThingName, Customer, HostName, data_dict, Installed)

                    # parse generic customer
                    abbr_dict_generic = indexed_config["generic"]
                    parse_configuration(abbr_dict_generic, temp_match_element, ThingName, Customer, HostName, data_dict, Installed)

            if log_line_count % 100000 == 0:
                print("records processed", log_line_count)
            log_line_count += 1
print("total records", log_line_count)

if len(customers_list_active) == 0:
    raise ValueError("number of recognized customers is zero")

customers_list_active = list(customers_list_active)
customers_list_active = sorted(customers_list_active)

# hde, solidcore, ens

for unit in data_dict:
    # add hde
    for entry in data_dict[unit]:
        if "HDE" in entry[1]:
            if data_dict[unit][entry] != NULL:
                data_dict[unit][(1000,"HDE")] = data_dict[unit][entry]
                break
    have_solidcore = "no"
    # add solidcore new
    for entry in data_dict[unit]:
        if "Solidcore" in entry[1]:
            if "New" in entry[1]:
                if data_dict[unit][entry] != NULL:
                    data_dict[unit][(1001,"Solidcore")] = data_dict[unit][entry]
                    have_solidcore = "yes"
                    break
    # add solidcore old
    if have_solidcore == "no":
        for entry in data_dict[unit]:
            if "Solidcore" in entry[1]:
                if data_dict[unit][entry] != NULL:
                    data_dict[unit][(1001,"Solidcore")] = data_dict[unit][entry]
                    break
    have_ens = "no"
    # add anti-virus new
    for entry in data_dict[unit]:
        if "Anti-Virus" in entry[1]:
            if "New" in entry[1]:
                if data_dict[unit][entry] != NULL:
                    data_dict[unit][(1002,"ENS")] = data_dict[unit][entry]
                    have_ens = "yes"
                    break
    # add anti-virus old
    if have_ens == "no":
        for entry in data_dict[unit]:
            if "Anti-Virus" in entry[1]:
                if data_dict[unit][entry] != NULL:
                    data_dict[unit][(1002,"ENS")] = data_dict[unit][entry]
                    break

# ssq

# ssq sort order reversed
sort_order_ssq_reversed = OrderedDict(reversed(list(sort_order_ssq.items())))
# ssq versioning parsing configuration
ssq_versions = OrderedDict()
for package in masterdict['generic']['package data']:
    for element in masterdict['generic']['package data'][package]:
        if "SSQ" in element:
            for name in element:
                if name != "SSQ":
                    if name not in ssq_versions:
                        ssq_versions[name]=[]
                    ssq_versions[name].append(package)
# reverse ordered dict
ssq_versions = OrderedDict(reversed(list(ssq_versions.items())))
# check and add ssq to data_dict
for unit in data_dict:
    disc_found_flag = 0
    for ssq in ssq_versions:
        if disc_found_flag == 0:
            disc_found_flag = 1
            for disc in ssq_versions[ssq]:
                if disc_found_flag == 0:
                    break
                elif disc_found_flag == 1:
                    disc_found_flag = 0
                    for package in data_dict[unit]:
                        if package[1] == disc:
                            disc_found_flag = 1
                            temp_ssq_name = ssq
                            break
        elif disc_found_flag == 1:
            # get number for tuple name
            num = sort_order_ssq[temp_ssq_name]
            # add info to unit
            data_dict[unit][(num, temp_ssq_name)] = data_dict[unit][package]
            # add other ssqs
            temp_flag = 0
            for ssq_in_order in sort_order_ssq_reversed:
                if temp_flag == 0:
                    if ssq_in_order != temp_ssq_name:
                        continue
                    else:
                        temp_flag = 1
                elif temp_flag == 1:
                    # get number
                    num = sort_order_ssq_reversed[ssq_in_order]
                    # add ssq in other columns
                    data_dict[unit][(num, ssq_in_order)]="Yes"    
            break
sort_order_ssq_reverse_key_value = {}
for element in sort_order_ssq:
    sort_order_ssq_reverse_key_value[(sort_order_ssq[element], element)] = None

# mssp

# get mssp sort order
sort_order_mssp = OrderedDict(masterdict["generic"]["package data"])
for key in list(sort_order_mssp):
    if not key.startswith("MSSP "):
        del sort_order_mssp[key]
sort_order_mssp_key = OrderedDict()
i = 0
for key in sort_order_mssp:
    sort_order_mssp_key[(3000 + i, key)] = sort_order_mssp[key][0][1]
    i = i + 1
temp_dict = {}
for unit in data_dict:
    for package in data_dict[unit]:
        for mssp in sort_order_mssp_key:
            if package[1] == mssp[1]:
                if unit not in temp_dict:
                    temp_dict[unit] = {}
                # generate key
                key = (sort_order_mssp_key[mssp])
                temp_dict[unit][key] = data_dict[unit][package]
# renumerate columns
temp_dict_ii = {}
for unit in temp_dict:
    if unit not in temp_dict_ii:
        temp_dict_ii[unit] = {}
    for package in temp_dict[unit]:
        number=list(sort_order_mssp_key.keys())[list(sort_order_mssp_key.values()).index(package)]
        package_tuple=(number[0],package)
        temp_dict_ii[unit][package_tuple]=temp_dict[unit][package]
# merge dictionaries
for temp_unit in temp_dict_ii:
    for package in temp_dict_ii[temp_unit]:
        data_dict[temp_unit][package] = temp_dict_ii[temp_unit][package]
# get audit mssp set
mssp_audit_set = set()
for mssp in sort_order_mssp_key:
    mssp_audit_set.add(sort_order_mssp_key[mssp])
sort_order_mssp_key_reversed = OrderedDict(reversed(list(sort_order_mssp_key.items())))
basic_mssp_order = {}
for mssp in sort_order_mssp_key_reversed:
    basic_mssp_order[sort_order_mssp_key_reversed[mssp]]=mssp[0]
basic_mssp_order_invert = []
for mssp in basic_mssp_order:
    basic_mssp_order_invert.append((basic_mssp_order[mssp], mssp))
basic_mssp_order_invert.sort(reverse=True)
# fill remaining fields for mssp with yes
for unit in data_dict:
    flag = 0
    for mssp in basic_mssp_order_invert:
        if flag == 0:
            if mssp in data_dict[unit]:
                flag = 1
        elif flag == 1:
            if mssp in data_dict[unit]:
                pass
            else:
                data_dict[unit][mssp] = "Yes"
  
# implement image
basic_mssp_order_invert.sort()
for unit in data_dict:
    if "image" in masterdict[unit[1]]:
        if "ssq" in masterdict[unit[1]]["image"]:
            ssq = masterdict[unit[1]]["image"]["ssq"]
            for ssq_temp in sort_order_ssq_reverse_key_value:
                if ssq_temp[1] != ssq:
                    data_dict[unit][ssq_temp] = None
                else:
                    break
        if "mssp" in masterdict[unit[1]]["image"]:
            mssp = masterdict[unit[1]]["image"]["mssp"]
            for mssp_temp in basic_mssp_order_invert:
                if mssp_temp[1] != mssp:
                    data_dict[unit][mssp_temp] = None
                else:
                    break

master_inventory_dataframe = pandas.DataFrame(data_dict)
master_inventory_dataframe = master_inventory_dataframe.transpose()
# sort columns by top numberd row
master_inventory_dataframe = master_inventory_dataframe.sort_index(axis=1)
# drop top numbered level
master_inventory_dataframe.columns = master_inventory_dataframe.columns.droplevel(0)
# set index
master_inventory_dataframe.index.names = ['ThingName', 'Customer', 'ATM ID']
# convert all 3 index levels to columns
master_inventory_dataframe.reset_index(inplace=True)
# rename columns
master_inventory_dataframe.rename(columns={"Customer": "Customer from inventory"}, inplace=True)
# cast all ids to upper
master_inventory_dataframe['ATM ID'] = master_inventory_dataframe['ATM ID'].apply(lambda x: x.upper())

if len(customers_list_active) == 0:
    raise ValueError("number of recognized customers is zero")
customers_list_active = list(customers_list_active)
customers_list_active = sorted(customers_list_active)

# audit

print("parsing through audits file")
for file in os.scandir(source_dir_audits):
    # print("parsing through file: " + file.name)
    audits_dataframes_dictionary = pandas.read_excel(file, sheet_name=None, engine='openpyxl')
del audits_dataframes_dictionary['Service CU offline list']
del audits_dataframes_dictionary['Sheet1']
master_audits_dataframe = pandas.concat(audits_dataframes_dictionary)
list_of_columns_in_dataframe_needed = [
    "ATM ID",
    "GROUP",
    "PILOTS",
    "UNIT IN LIFECYCLE",
    "CD-CashDisp / DA-Deposit",
    "Wired/Wireless",
    "Zone",
    "WIN",
    "BASE SOFTWARE",
    "CAPTURE",
    "CAPTURE ITM",
    "Apple V6",
    "Apple CSS",
    "Apple Qualys",
    "Apple ServicePath",
    "Apple DBX",
    "Fremont Settlement Fis Host",
    "NDC Personalization",
    "HOTFIX V891",
    "HOTFIX W1492",
    "UA",
    "RBU",
    "WPM",
    "HDE",
    "Solidcore",
    "ENS",
    "SSQ 19",
    #"SSQ1 19",
    #"SSQ2 19",
    #"SSQ3 19",
    #"SSQ4 19",
    "SSQ1 20",
    "SSQ2 20",
    "SSQ3 20",
    #"HOTFIX W1242",
    "SSQ4 20",
    "SSQ4.1 20",
    "SSQ1 21",
    "SSQ1 21 T",
    "SSQ2 21",
    "SSQ3 21",
    "SSQ4 21",
    "SSQ4.1 21",
    "SSQ1 22",
    "AE 3.5.1.453 Hotfix 10 PSSCM-15667",
    #"MSSP Q1 18",
    #"MSSP Q2 18",
    #"MSSP Q3 18",
    #"MSSP Q4 18",
    "MSSP Q1 19",
    "MSSP Q2 19",
    "MSSP Q3 19",
    "MSSP Q4 19",
    #"MSSP Q4 19 + JAN 20",
    #"ESU",
    "MSSP Q1 20",
    "MSSP Q2 20",
    "MSSP Q3 20",
    "MSSP Q4 20",
    #"MSSP JAN 21",
    "MSSP Q1 21",
    "MSSP Q2 21",
    "MSSP Q3 21",
    "MSSP Q4 21",
    "MSSP Q1 22",
    "SWD Notes",
    "SWD Notes2",
    "GROUPS",
    "PLAN"
]
# this one is needed below also so redesignation is needed so that we don't have name overwrite
list_of_columns_in_dataframe_needed_ii = list_of_columns_in_dataframe_needed
# order columns
try:
    master_audits_dataframe = master_audits_dataframe[list_of_columns_in_dataframe_needed]
except Exception as ex:
    print(ex)
    sleep(999)
# set index
master_audits_dataframe.index.names = ['Customer', 'Index']
# convert all index levels to columns
master_audits_dataframe.reset_index(inplace=True)
# rename columns
master_audits_dataframe.rename(columns={"Customer": "Customer from audits"}, inplace=True)
# cast all ids to number or string ? or all to string ?
master_audits_dataframe['ATM ID'] = master_audits_dataframe['ATM ID'].apply(lambda x: x.upper() if (not isinstance(x, int) and not isinstance(x, float)) else x)
master_audits_dataframe['ATM ID'] = master_audits_dataframe['ATM ID'].apply(lambda x: str(x) if not isinstance(x, str) else x)

# imacs

print("parsing through imacs files")
list_of_imacs_dataframes = []
list_of_columns_in_dataframe_needed = [
    "Customer Name",
    "Customer Prefix",
    "Computer Name",
    "ATM ID",
    "Time Zone",
    "SST IP Address",
    # "Application SW",
    # "PC Core",
    "SST Function Type",
    "ATM/ITM/ITM as ATM"
]
for subdir, dirs, files in os.walk(source_dir_imacs):
    for filename in files:
        filepath = subdir + os.sep + filename
        file_extension = os.path.splitext(filename)[1]
        if file_extension == '.xlsx':
            dataframes = pandas.read_excel(filepath, sheet_name=None, engine='openpyxl')
            for dataframe in dataframes:
                # print("parsing file - dataframe ", filepath, " - ", dataframe)
                
                # trim spaces from all data
                dataframes[dataframe] = dataframes[dataframe].applymap(lambda x: x.strip() if isinstance(x, str) else x)

                # get/set number of rows and columns

                num_of_rows = len(dataframes[dataframe])
                num_of_columns = len(dataframes[dataframe].columns)

                loop_rows = 30
                loop_columns = 30

                if num_of_rows < loop_rows:
                    loop_rows = num_of_rows
                if num_of_columns < loop_columns:
                    loop_columns = num_of_columns

                # loop through data and drop rows
                flag = 0
                for i in range(loop_rows):
                    for j in range(loop_columns):
                        value = dataframes[dataframe].iloc[0,j]
                        value = str(value)
                        if "Customer Name" in value:
                            flag = 1
                            break
                    if flag == 1:
                        break
                    dataframes[dataframe].drop(axis=0, index=dataframes[dataframe].index[0], inplace=True)

                # check number of rows
                num_of_rows = len(dataframes[dataframe])
                if num_of_rows < 2:
                    continue

                # set header row and remove duplicate first row
                dataframes[dataframe].columns = dataframes[dataframe].iloc[0]
                dataframes[dataframe] = dataframes[dataframe][1:]

                # rename column label
                if "Computer Name" in dataframes[dataframe].columns:
                    dataframes[dataframe].rename(columns={"Computer Name": "ATM ID"}, inplace=True)
                if "Computer Name/ Current ID" in dataframes[dataframe].columns:
                    dataframes[dataframe].rename(columns={"Computer Name/ Current ID": "ATM ID"}, inplace=True)

                # drop nan column labels
                dataframes[dataframe].columns = dataframes[dataframe].columns.fillna('to_drop')
                if "to_drop" in dataframes[dataframe].columns:
                    dataframes[dataframe].drop('to_drop', axis = 1, inplace = True)

                # drop not needed columns
                list_of_columns_in_dataframe_all = list(dataframes[dataframe].columns)
                list_of_columns_in_dataframe_to_drop = []
                for item in list_of_columns_in_dataframe_all:
                    if item not in list_of_columns_in_dataframe_needed:
                        list_of_columns_in_dataframe_to_drop.append(item)
                for column in list_of_columns_in_dataframe_to_drop:
                        dataframes[dataframe].drop(column, axis=1, inplace=True)

                # check for duplicate column names
                list_of_columns = dataframes[dataframe].columns
                num_of_columns = len(list_of_columns)
                num_of_unique_columns = len(set(list_of_columns))
                if num_of_columns != num_of_unique_columns:
                    raise Exception("duplicate column in imac ", filepath, " - ", dataframe)

                # create master list of imacs dataframes
                list_of_imacs_dataframes.append(dataframes[dataframe])

# concat all imacs dataframes
master_imacs_dataframe = pandas.concat(list_of_imacs_dataframes, axis=0, ignore_index=True)
# drop missing unit id values
master_imacs_dataframe.dropna(subset=["ATM ID"], inplace=True)
# rename columns
master_imacs_dataframe.rename(columns={"Customer Name": "Customer from imacs"}, inplace=True)
# cast all ids to number or string ? or all to string ?
master_imacs_dataframe['ATM ID'] = master_imacs_dataframe['ATM ID'].apply(lambda x: x.upper() if not isinstance(x, int) else x)
master_imacs_dataframe['ATM ID'] = master_imacs_dataframe['ATM ID'].apply(lambda x: str(x) if not isinstance(x, str) else x)
# unify timezones
for index, row in master_imacs_dataframe.iterrows():
    for zone in zones:
        for element in zones[zone]:
            if element == row['Time Zone']:
                master_imacs_dataframe.at[index, 'Time Zone'] = zone

audits_duplicates_dataframe = master_audits_dataframe[master_audits_dataframe.duplicated('ATM ID')]
if not audits_duplicates_dataframe.empty:
    print()
    print("duplicates in audits")
    print(audits_duplicates_dataframe)
    raise Exception("check for duplicates in audits")
else:
    print("no duplicates in audits")

imacs_duplicates_dataframe = master_imacs_dataframe[master_imacs_dataframe.duplicated('ATM ID')]
if not imacs_duplicates_dataframe.empty:
    print()
    print("duplicates in imacs")
    print(imacs_duplicates_dataframe)
    print("~and~")
    print(imacs_duplicates_dataframe.to_string())
    raise Exception("check for duplicates in imacs")
else:
    print("no duplicates in imacs")

print("generating master dataframe")
master_dataframe_list = [
    master_devices_dataframe,
    master_agents_dataframe,
    master_imacs_dataframe,
    master_audits_dataframe,
    master_inventory_dataframe
]
master_master_dataframe = reduce(lambda left, right: pandas.merge(left, right, on=["ATM ID"], how="outer"), master_dataframe_list)
# reduced number of ids
master_master_dataframe_filtered = master_master_dataframe.dropna(subset=["name", "Customer from audits"], how='all')

print("analyzing data")

duplicate_atm_id = master_master_dataframe_filtered[master_master_dataframe_filtered.duplicated('ATM ID')]
if not duplicate_atm_id.empty:
    # print()
    # print("duplicate atm id")
    # print(duplicate_atm_id)
    with open(reports_dir + os.sep + 'analysis.csv', mode='a') as file:
        writer = csv.writer(file)
        writer.writerow([""])
        writer.writerow(["duplicate atm id",""])
    duplicate_atm_id.to_csv(reports_dir + os.sep + 'analysis.csv', mode='a')

duplicate_ip_address = master_master_dataframe_filtered[master_master_dataframe_filtered.duplicated('Asset_IP')]
duplicate_ip_address = duplicate_ip_address.dropna(subset=['Asset_IP'])
if not duplicate_ip_address.empty:
    # print()
    # print("duplicate ip address")
    # print(duplicate_ip_address)
    with open(reports_dir + os.sep + 'analysis.csv', mode='a') as file:
        writer = csv.writer(file)
        writer.writerow([""])
        writer.writerow(["duplicate ip address",""])
    duplicate_ip_address.to_csv(reports_dir + os.sep + 'analysis.csv', mode='a')

offline_units = master_master_dataframe_filtered[master_master_dataframe_filtered['isConnected'] == False]
if not offline_units.empty:
    # print()
    # print("offline units")
    # print(offline_units)
    with open(reports_dir + os.sep + 'analysis.csv', mode='a') as file:
        writer = csv.writer(file)
        writer.writerow([""])
        writer.writerow(["offline units",""])
    offline_units.to_csv(reports_dir + os.sep + 'analysis.csv', mode='a')

compare_atmIDs_dcs_audits = master_master_dataframe_filtered[((master_master_dataframe_filtered["Customer from devices"].isna()) | (master_master_dataframe_filtered["Customer from audits"].isna()))]
if not compare_atmIDs_dcs_audits.empty:
    # print()
    # print("compare atm ids for dcs & audits")
    # print(compare_atmIDs_dcs_audits)
    with open(reports_dir + os.sep + 'analysis.csv', mode='a') as file:
        writer = csv.writer(file)
        writer.writerow([""])
        writer.writerow(["compare atm ids for dcs & audits",""])
    compare_atmIDs_dcs_audits.to_csv(reports_dir + os.sep + 'analysis.csv', mode='a')

compare_ipAdresses_dcs_imacs = master_master_dataframe_filtered[master_master_dataframe_filtered["Asset_IP"] != master_master_dataframe_filtered["SST IP Address"]]
if not compare_ipAdresses_dcs_imacs.empty:
    # print()
    # print("compare ip adresses dcs & imacs")
    # print(compare_ipAdresses_dcs_imacs)
    with open(reports_dir + os.sep + 'analysis.csv', mode='a') as file:
        writer = csv.writer(file)
        writer.writerow([""])
        writer.writerow(["compare ip adresses dcs & imacs",""])
    compare_ipAdresses_dcs_imacs.to_csv(reports_dir + os.sep + 'analysis.csv', mode='a')

compare_zones_dcs_imacs_audit = master_master_dataframe_filtered[(master_master_dataframe_filtered["Zone"] != master_master_dataframe_filtered["Time Zone"]) & (master_master_dataframe_filtered["Time Zone"] != master_master_dataframe_filtered["timeZoneOffset"]) & (master_master_dataframe_filtered["Zone"] != master_master_dataframe_filtered["timeZoneOffset"])]
if not compare_zones_dcs_imacs_audit.empty:
    # print()
    # print("compare zones dcs & imacs & audit")
    # print(compare_zones_dcs_imacs_audit)
    with open(reports_dir + os.sep + 'analysis.csv', mode='a') as file:
        writer = csv.writer(file)
        writer.writerow([""])
        writer.writerow(["compare zones dcs & imacs & audit",""])
    compare_zones_dcs_imacs_audit.to_csv(reports_dir + os.sep + 'analysis.csv', mode='a')

# with inverse selection ~
compare_windows_mssp_version = master_master_dataframe_filtered[~((master_master_dataframe_filtered["WIN"] == "2016 LTSB 1607") & (master_master_dataframe_filtered["osVersion"] == "Microsoft Windows 10 Enterprise 2016 LTSB"))]
if not compare_windows_mssp_version.empty:
    # print()
    # print("compare microsoft mssp version")
    # print(compare_windows_mssp_version)
    with open(reports_dir + os.sep + 'analysis.csv', mode='a') as file:
        writer = csv.writer(file)
        writer.writerow([""])
        writer.writerow(["compare microsoft mssp version",""])
    compare_windows_mssp_version.to_csv(reports_dir + os.sep + 'analysis.csv', mode='a')

# with inverse selection ~
compare_base_core_versions = master_master_dataframe_filtered[
    ~((
        (master_master_dataframe_filtered["BASE SOFTWARE"] == "EDGE 10") & 
        ((master_master_dataframe_filtered["TYPE/CORE"] == "10.01.03.13") |
        (master_master_dataframe_filtered["TYPE/CORE"] == "10.01.04.4") |
        (master_master_dataframe_filtered["TYPE/CORE"] == "10.01.06.1"))
    )
    |
    (
        (master_master_dataframe_filtered["BASE SOFTWARE"] == "ITM") &
        ((master_master_dataframe_filtered["TYPE/CORE"] == "02.00.02.1") |
        (master_master_dataframe_filtered["TYPE/CORE"] == "02.00.00.21") |
        (master_master_dataframe_filtered["TYPE/CORE"] == "02.00.01.4") |
        (master_master_dataframe_filtered["TYPE/CORE"] == "01.05.01.7"))
    ))
]
if not compare_base_core_versions.empty:
    # print()
    # print("compare base and core versions")
    # print(compare_base_core_versions)
    with open(reports_dir + os.sep + 'analysis.csv', mode='a') as file:
        writer = csv.writer(file)
        writer.writerow([""])
        writer.writerow(["compare base and core versions",""])
    compare_base_core_versions.to_csv(reports_dir + os.sep + 'analysis.csv', mode='a')

# missmatched elements dcs and audit
list_of_doubled_columns = []
for column_name in master_master_dataframe_filtered:
    if column_name.endswith("_x"):
        column_name_new = column_name.removesuffix("_x")
        list_of_doubled_columns.append(column_name_new)
for column in list_of_doubled_columns:
    column_x = column + "_x"
    column_y = column + "_y"
    missmatched_elements = master_master_dataframe_filtered[((master_master_dataframe_filtered[column_x].isnull() & master_master_dataframe_filtered[column_y].notnull()) | (master_master_dataframe_filtered[column_x].notnull() & master_master_dataframe_filtered[column_y].isnull()))]
    if not missmatched_elements.empty:
        # print()
        # print("missmatched elements dcs and audit, column: ", column)
        # print(missmatched_elements)
        with open(reports_dir + os.sep + 'analysis.csv', mode='a') as file:
                writer = csv.writer(file)
                writer.writerow([""])
                writer.writerow(["missmatched elements dcs and audit, column: ", column])
        missmatched_elements.to_csv(reports_dir + os.sep + 'analysis.csv', mode='a')

print("generating primary reports")
master_devices_dataframe.to_excel(reports_dir_primary + os.sep + 'master_devices_dataframe.xlsx')
master_agents_dataframe.to_excel(reports_dir_primary + os.sep + 'master_agents_dataframe.xlsx')
master_inventory_dataframe.to_excel(reports_dir_primary + os.sep + 'master_inventory_dataframe.xlsx')
master_audits_dataframe.to_excel(reports_dir_primary + os.sep + 'master_audits_dataframe.xlsx')
master_imacs_dataframe.to_excel(reports_dir_primary + os.sep + 'master_imacs_dataframe.xlsx')

print("generating master reports")
master_master_dataframe.to_excel(reports_dir + os.sep + 'master_master_dataframe.xlsx')
master_master_dataframe_filtered.to_excel(reports_dir + os.sep + 'master_master_dataframe_filtered.xlsx')

print("generating single customer reports")
for customer in customers_list_active:
    single_customer = master_master_dataframe_filtered[master_master_dataframe_filtered["Customer from devices"] == customer]
    single_customer = single_customer.dropna(how='all', axis='columns')
    single_customer.to_excel(reports_dir_single + os.sep + customer + '.xlsx')

print("generating audits")
list_to_drop = []
for column in master_master_dataframe_filtered:
    if column.endswith("_x"):
        list_to_drop.append(column)
master_audit_dataframe_filtered = master_master_dataframe_filtered.drop(list_to_drop, axis = 1)
master_audit_dataframe_filtered.columns = master_audit_dataframe_filtered.columns.str.replace('_y', '')
list_of_columns_in_dataframe_needed_ii.insert(0, "Customer")
master_audit_dataframe_filtered.rename(columns = {'Customer from devices':'Customer'}, inplace = True)
master_audit_dataframe_filtered = master_audit_dataframe_filtered.drop(columns=[col for col in master_audit_dataframe_filtered if col not in list_of_columns_in_dataframe_needed_ii])
master_audit_dataframe_filtered = master_audit_dataframe_filtered[list_of_columns_in_dataframe_needed_ii]
master_audit_dataframe_filtered.index.names = ['index']
master_audit_dataframe_filtered = master_audit_dataframe_filtered.set_index("ATM ID")
master_audit_dataframe_filtered = master_audit_dataframe_filtered.drop(["SWD Notes", "SWD Notes2", "GROUPS", "PLAN", 'index'], axis = 1, errors = 'ignore')
for customer in customers_list_active:
    single_customer = master_audit_dataframe_filtered[master_audit_dataframe_filtered["Customer"] == customer]
    single_customer = single_customer.dropna(how='all', axis='columns')
    single_customer = single_customer.drop("Customer", axis = 1)
    single_customer.to_excel(reports_dir_audits + os.sep + customer + '.xlsx')

print("formatting excel reports")
redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
greenFill = PatternFill(start_color='FFccffcc', end_color='FFccffcc', fill_type='solid')
date_style = NamedStyle(name='date_style', number_format='MM/DD/YYYY')
fill_cap = PatternFill(fill_type='solid', start_color='2e6990', end_color='2e6990')
font_cap = Font(bold=True, color='FFFFFFFF')
alignment=Alignment(horizontal='center')
border = Border(
    left=Side(border_style='thin', color='FF000000'),
    right=Side(border_style='thin', color='FF000000'),
    top=Side(border_style='thin', color='FF000000'),
    bottom=Side(border_style='thin', color='FF000000'),
)
for subdir, dirs, files in os.walk(reports_dir):
    for filename in files:
        print("formatting file: ", filename)
        filepath = subdir + os.sep + filename
        if filename == "analysis.csv":
            continue
        workbook = openpyxl.load_workbook(filepath)
        worksheet = workbook.active

        # style for dates
        workbook.add_named_style(date_style)

        # insert table
        table_name = worksheet.title.replace(" ", "")
        table_length = worksheet.max_row
        table_width = get_column_letter(worksheet.max_column)
        tab = openpyxl.worksheet.table.Table(displayName=table_name, ref=f'A1:{table_width}{table_length}')
        style = openpyxl.worksheet.table.TableStyleInfo(name="None", showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        worksheet.add_table(tab)

        # autofit column width
        for column in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column)
            length = length + 3
            worksheet.column_dimensions[column[0].column_letter].width = length

        # center justify & style dates
        for column in worksheet.iter_cols():
            for cell in column:
                cell.alignment = Alignment(horizontal='center')
                if isinstance(cell, datetime.datetime):
                    cell.style = date_style

        if subdir == r"reports\audit reports":

            # format for all cells
            for row in worksheet.rows:
                for cell in row:
                    cell.border = border
                    if cell.internal_value == "Yes":
                        cell.fill = greenFill
                    if cell.is_date:
                        cell.fill = greenFill

            # format for first row
            first_row = list(worksheet.rows)[0]
            for cell in first_row:
                cell.fill = fill_cap
                cell.font = font_cap
                
        else:

            # mark duplicates
            temp_cell = worksheet.cell(1,1)
            for cell in worksheet['D']:
                if cell.internal_value == temp_cell.internal_value:
                    cell.fill = redFill
                    temp_cell.fill = redFill
                temp_cell = cell

            # freeze panes
            cell = worksheet['E2']
            worksheet.freeze_panes = cell

        workbook.save(filepath)
print()
print("all reports finished")
sleep(2)
