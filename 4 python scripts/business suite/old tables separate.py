
# libraries needed for exporting to excel
# pip install openpyxl

import os
import shutil
from copy import copy
from time import sleep
import openpyxl.worksheet.table
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

print("generating directory structure")

reports_dir = "reports"
source_dir = "source"

try:
    shutil.rmtree(reports_dir)
except:
    print("error deleting reports directory structure")
    sleep(3)

# wait for windows to finish
sleep(3)
os.mkdir(reports_dir)

print("parsing source")

for subdir, dirs, files in os.walk(source_dir):
    for filename in files:
        filepath = subdir + os.sep + filename
        master_workbook = openpyxl.load_workbook(filepath)

        for i in range(len(master_workbook.worksheets)):
            old_sheet = master_workbook.worksheets[i]

            # get data from sheet
            customer_name = old_sheet.title
            customer_file_path = reports_dir + os.sep + customer_name + " fleet.xlsx"
            max_row = old_sheet.max_row
            max_column = old_sheet.max_column

            print("reading customer ", customer_name, ", dim: ", max_row, " x ", max_column)

            # create new workbook & sheet
            new_workbook = Workbook()
            new_sheet = new_workbook.active

            # # copying data
            for i in range(1, max_row + 1):
                for j in range(1, max_column + 1 - 3):  # -4 for removing 4 columns from the end
                    old_cell = old_sheet.cell(row=i, column=j)
                    new_cell = new_sheet.cell(row=i, column=j)
                    new_cell.value = old_cell.value

                    # styles
                    new_cell.alignment = Alignment(horizontal='center')
                    new_cell.number_format = copy(old_cell.number_format)
                    new_cell.fill = copy(old_cell.fill)
                    new_cell.font = copy(old_cell.font)
                    new_cell.border = copy(old_cell.border)

            # autofit column width
            for col in new_sheet.columns:
                for cell in col:
                    new_sheet.column_dimensions[get_column_letter(cell.column)].auto_size = True
                    break

            new_workbook.save(filename=customer_file_path)
print()
print("all reports finished")
sleep(3)
